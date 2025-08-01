import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

st.title("Proposal Preferences Formatter")

uploaded_files = st.file_uploader(
    "Upload preference files (multiple allowed)",
    type=["xlsx"],
    accept_multiple_files=True
)

template_file = st.file_uploader(
    "Upload the formatting template file (optional)",
    type=["xlsx"]
)

if st.button("Go"):
    if not uploaded_files and not template_file:
        st.error("Please upload at least one preference file or a formatting template.")
    else:
        # Process preference files
        if uploaded_files:
            st.write(f"Processing {len(uploaded_files)} preference file(s)")
            for uploaded_file in uploaded_files:
                st.write(f"Processing file: {uploaded_file.name}")
                try:
                    uploaded_file.seek(0)
                    xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
                    st.write(f"Sheets found: {xls.sheet_names}")
                    if len(xls.sheet_names) == 0:
                        st.error(f"File '{uploaded_file.name}' has no sheets.")
                        continue
                    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
                    st.write(f"Preview of {uploaded_file.name}:")
                    st.dataframe(df.head())
                except Exception as e:
                    st.warning(f"Pandas failed to read '{uploaded_file.name}': {e}")
                    try:
                        uploaded_file.seek(0)
                        wb = load_workbook(filename=io.BytesIO(uploaded_file.read()))
                        st.write(f"openpyxl sheets: {wb.sheetnames}")
                        if not wb.sheetnames:
                            st.error(f"openpyxl found no sheets in '{uploaded_file.name}'.")
                            continue
                        ws = wb[wb.sheetnames[0]]
                        data = ws.values
                        cols = next(data)
                        data = list(data)
                        df = pd.DataFrame(data, columns=cols)
                        st.write(df.head())
                    except Exception as e2:
                        st.error(f"openpyxl also failed for '{uploaded_file.name}': {e2}")

        # Process template file if provided
        if template_file:
            st.write("Processing formatting template file...")
            try:
                template_file.seek(0)
                xls = pd.ExcelFile(template_file, engine='openpyxl')
                st.write(f"Template sheets: {xls.sheet_names}")
                if len(xls.sheet_names) == 0:
                    st.error("Template file has no sheets.")
                else:
                    df_template = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
                    st.write(df_template.head())
            except Exception as e:
                st.error(f"Failed to read template file: {e}")
